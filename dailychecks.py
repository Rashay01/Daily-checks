from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from dotenv import load_dotenv
import os
import time
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime

load_dotenv()


def get_current_day_in_words():
    # Get the current date
    now = datetime.now()

    # Get the full name of the current day
    day_name = now.strftime("%A")

    return day_name


def bypass_security_warning(driver):
    try:
        # Wait for 50 seconds before attempting to click "Advanced"
        # time.sleep(50)

        # Wait for the "Advanced" button to be clickable and click it
        advanced_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//*[text()="Advanced"]'))
        )
        advanced_button.click()

        # Wait for the "Proceed to [site] (unsafe)" link to be clickable and click it
        proceed_link = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable(
                (By.XPATH, '//*[text()="Proceed to [site] (unsafe)"]')
            )
        )
        proceed_link.click()

        print("Bypassed security warning and proceeded to the site.")

    except Exception as e:
        print(f"An error occurred while bypassing the security warning: {e}")


def click_element_by_text(driver, text):
    try:
        # Wait for any element with the specified text to be clickable
        element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, f'//*[text()="{text}"]'))
        )

        # Click the element
        element.click()

        # Wait for new tab to open (optional: you can adjust this based on how long it typically takes)
        time.sleep(2)

        # Switch to the new tab
        driver.switch_to.window(driver.window_handles[-1])

        # Wait for the page to load fully (adjust as needed)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*"))
        )

        # # Take a screenshot
        # screenshot_path = f'screenshot_{text}.png'
        # driver.save_screenshot(screenshot_path)
        # print(f'Screenshot of the new tab saved as {screenshot_path}.')

    except Exception as e:
        print(f"An error on click: {text}")
        raise


def click_back_icon(driver):
    try:
        # Wait for the element with the specified class and style to be clickable
        element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable(
                (By.CSS_SELECTOR, "span.sapUshellShellHeadItmCntnt")
            )
        )

        # Click the element
        element.click()

        # Wait for new tab to open (optional: you can adjust this based on how long it typically takes)
        time.sleep(2)

        # Switch to the new tab (if a new tab is opened)
        if len(driver.window_handles) > 1:
            driver.switch_to.window(driver.window_handles[-1])

            # Wait for the page to load fully (adjust as needed)
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//*"))
            )

            # Optionally, take a screenshot
            # screenshot_path = 'screenshot_new_tab.png'
            # driver.save_screenshot(screenshot_path)
            # print(f'Screenshot of the new tab saved as {screenshot_path}.')

    except Exception as e:
        print(f"An error occurred while clicking the element: {e}")
        raise


def click_element_by_id(driver, elementID, wait_for_element_xpath=None):
    try:
        # Wait for the element with the specified ID to be clickable
        element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, elementID))
        )

        # Click the element
        element.click()

        # Wait for a new tab to open
        WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(2))

        # Switch to the new tab
        driver.switch_to.window(driver.window_handles[-1])

        # Wait for the page to load fully or for a specific element to appear
        if wait_for_element_xpath:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, wait_for_element_xpath))
            )
        else:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//*"))
            )

        # Optional: Take a screenshot
        # screenshot_path = f'screenshot_{elementID}.png'
        # driver.save_screenshot(screenshot_path)
        # print(f'Screenshot of the new tab saved as {screenshot_path}.')

    except Exception as e:
        print(f"An error occurred while clicking element with ID '{elementID}': {e}")
        raise


def click_element_by_text_js(driver, text, wait_time=10):
    """
    Clicks an element with the specified text using JavaScript.

    :param driver: Selenium WebDriver instance.
    :param text: Text of the element to be clicked.
    :param wait_time: Maximum time to wait for the page to load.
    """
    try:
        # Wait for the page to load
        WebDriverWait(driver, wait_time).until(
            EC.presence_of_element_located((By.XPATH, "//*"))
        )

        # Use JavaScript to find and click the element
        script = f"""
        var elements = document.querySelectorAll('a');
        for (var i = 0; i < elements.length; i++) {{
            if (elements[i].innerText.includes("{text}")) {{
                elements[i].click();
                return;
            }}
        }}
        """
        driver.execute_script(script)
        print(f"Element with text '{text}' clicked successfully using JavaScript.")
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*"))
        )
    except Exception as e:
        print(f"An error occurred while clicking {text} with JavaScript: {e}")
        raise


def click_element_when_ready(driver, text, wait_time=10):
    """
    Waits for an element with the specified text to be clickable and clicks it.

    :param driver: Selenium WebDriver instance.
    :param text: Text of the element to be clicked.
    :param wait_time: Maximum time to wait for the element to be clickable.
    """
    try:
        # Wait for the element to be clickable
        element = WebDriverWait(driver, wait_time).until(
            EC.element_to_be_clickable((By.XPATH, f'//*[text()="{text}"]'))
        )
        element.click()
        print(f"Element with text '{text}' clicked successfully.")
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*"))
        )
    except Exception as e:
        print(f"An error occurred while waiting for and clicking {text}: {e}")
        raise


def scroll_and_click_by_text(
    driver, text, wait_time=10, max_scroll_attempts=20, scroll_increment=1
):
    """
    Scrolls down the page incrementally until an element with the specified text is found and clicks it.

    :param driver: Selenium WebDriver instance.
    :param text: Text of the element to be clicked.
    :param wait_time: Maximum time to wait for the element to be present and clickable.
    :param max_scroll_attempts: Maximum number of incremental scrolls before giving up.
    :param scroll_increment: Amount of pixels to scroll per attempt.
    """
    try:
        # Initialize WebDriverWait
        wait = WebDriverWait(driver, wait_time)

        for attempt in range(max_scroll_attempts):
            # Wait for the element to be present
            try:
                element = wait.until(
                    EC.presence_of_element_located((By.XPATH, f'//*[text()="{text}"]'))
                )
                # Scroll to the element using JavaScript
                driver.execute_script("arguments[0].scrollIntoView(true);", element)

                # Optional: Wait a bit to ensure the scrolling is done
                time.sleep(5)

                # Click the element
                element.click()
                print(f"Element with text '{text}' clicked successfully.")
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, "//*"))
                )
                return  # Exit the function after successful click
            except:
                # If the element is not found, scroll down incrementally and try again
                driver.execute_script(f"window.scrollBy(0, {scroll_increment});")
                time.sleep(2)  # Wait for the page to load new content

        print(f"Element with text '{text}' not found after scrolling.")
    except Exception as e:
        print(f"An error occurred while scrolling and clicking {text}: {e}")
        raise


def click_image_by_id(driver, image_id, wait_time=10):
    try:
        # Wait for the image element to be present
        wait = WebDriverWait(driver, wait_time)
        image_element = wait.until(EC.presence_of_element_located((By.ID, image_id)))

        # Scroll to the image element using JavaScript
        driver.execute_script("arguments[0].scrollIntoView(true);", image_element)

        # Optional: Wait a bit to ensure the scrolling is done
        WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.ID, image_id)))

        # Click the image element
        image_element.click()
        print(f"Image with ID '{image_id}' clicked successfully.")
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*"))
        )

    except Exception as e:
        print(f"An error occurred in clicking image: {image_id}")
        raise


def click_button_by_value(driver, value):
    try:
        # Locate the button using the 'value' attribute
        button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable(
                (By.XPATH, f'//input[@type="submit" and @value="{value}"]')
            )
        )

        # Click the button
        button.click()

        # Wait for new tab to open (optional: you can adjust this based on how long it typically takes)
        time.sleep(2)

        # Switch to the new tab if applicable
        if len(driver.window_handles) > 1:
            driver.switch_to.window(driver.window_handles[-1])

        # Wait for the page to load fully (adjust as needed)
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*"))
        )

        print(f'Button with value "{value}" clicked.')

    except Exception as e:
        print(f"Error occurred while clicking the button with value '{value}'")
        # Reraise the exception to indicate failure
        raise


def click_element_by_partial_id(driver, partial_id):
    try:
        # Wait for the element with the partial ID to be clickable
        element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable(
                (By.XPATH, f"//*[contains(@id, '{partial_id}')]")
            )
        )

        # Click the element
        element.click()

        # Optionally wait for a new tab to open (adjust the sleep time as needed)
        time.sleep(2)

        # Switch to the new tab (if applicable)
        if len(driver.window_handles) > 1:
            driver.switch_to.window(driver.window_handles[-1])

        # Optionally wait for the page to load fully
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*"))
        )

    except Exception as e:
        print(
            f"An error occurred while clicking the element with partial ID '{partial_id}': {e}"
        )
        raise


def get_text_by_partial_id(driver, partial_id):
    try:
        # Wait for the element with the partial ID to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.XPATH, f"//*[contains(@id, '{partial_id}')]")
            )
        )

        # Get the text from the element
        text = element.text
        return text

    except Exception as e:
        print(
            f"An error occurred while retrieving text from the element with partial ID '{partial_id}': {e}"
        )
        return None


# -------------------------------------------------------------------------------------
def selectDatabase(driver, groupName, databaseName, fileLocation):
    try:
        click_image_by_id(driver, "shell-header-icon")
        print(f"Click group {groupName}")
        click_element_by_text(driver, groupName)

        time.sleep(2)

        print(f"Scroll and click {databaseName}")
        click_element_by_text_js(driver, databaseName)

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*"))
        )
        time.sleep(20)

        print(f"Successfully selected database")
        save_screenshot(driver, fileLocation, "databaseSelected.png")
        save_screenshot()
        print("databaseSelected Screenshot saved.")

    except Exception as e:
        print(f"Error occurred in Select Database")


def login(driver):
    try:
        # Load username and password from environment variables
        username = os.getenv("USERNAME_2")
        password = os.getenv("PASSWORD")

        # Locate the username and password input fields
        username_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.NAME, "username")
            )  # Adjust selector as needed
        )
        password_field = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located(
                (By.NAME, "password")
            )  # Adjust selector as needed
        )

        # Input the username and password
        username_field.send_keys(username)
        password_field.send_keys(password)

        # Submit the form or perform login action
        click_button_by_value(driver, "Log On")

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*"))
        )
        # time.sleep(4)
        wait = WebDriverWait(driver, 10)
        image_element = wait.until(
            EC.presence_of_element_located((By.ID, "shell-header-icon"))
        )

        # Check if the text matches "My Home"
        if image_element:
            print('The text "My Home" is present in the div.')
            print("Logged in successfully.")
            return 1
        else:
            print('The text "My Home" is not present in the div.')
            return -1

    except Exception as e:
        print(f"An error occurred during login: {e}")
        return -1


def check_and_click_elements(driver, wait_time=10):
    try:
        # Wait for the presence of the text elements
        wait = WebDriverWait(driver, wait_time)

        # Check for the presence of the "Monitoring" text
        monitoring_element = wait.until(
            EC.presence_of_element_located((By.XPATH, '//span[text()="Monitoring"]'))
        )

        # If "Monitoring" is found, click the associated button
        if monitoring_element:
            print("Text 'Monitoring' found.")
            # Locate and click the button associated with "Monitoring"
            #
            click_element_by_id(driver, "__xmlview38--idVariantManagement-trigger-img")
            # button = driver.find_element(By.XPATH, '//span[@id="__xmlview188--idVariantManagement-trigger-inner"]')
            # button.click()
            print("Button associated with 'Monitoring' clicked.")

        # Check for the presence of the "All" text

        # If "All" is found, click the associated element

        # click_element_by_text(driver, "All")
        click_element_by_id(driver, "__xmlview38--idVariantManagement-trigger-item-1")
        print("Element with text 'All' clicked.")
        return True

    except Exception as e:
        print(f"An error occurred: {e}")
        raise


def search_input_field(driver, input_id, search_query, wait_for_element_xpath=None):
    """
    Searches using an <input> field identified by its ID and optionally waits for a specific element to appear.

    :param driver: The Selenium WebDriver instance.
    :param input_id: The ID of the <input> element used for search.
    :param search_query: The query string to send to the search input field.
    :param wait_for_element_xpath: Optional XPath of an element to wait for after performing the search.
    """
    try:
        # Wait for the <input> element with the specified ID to be visible and enabled
        input_field = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.ID, input_id))
        )

        # Clear the input field (in case it contains any pre-existing value)
        input_field.clear()

        # Send the search query to the input field
        input_field.send_keys(search_query)

        # Optionally wait for a specific element or condition if provided
        if wait_for_element_xpath:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, wait_for_element_xpath))
            )

        # Optional: Take a screenshot
        # screenshot_path = f'screenshot_search_{search_query}.png'
        # driver.save_screenshot(screenshot_path)
        # print(f'Screenshot saved as {screenshot_path}.')
        return True

    except Exception as e:
        print(f"An error occurred while searching with input ID '{input_id}': {e}")
        raise


def save_screenshot(driver, folder_path, file_name):
    """
    Takes a screenshot and saves it to a specified folder with a given file name.

    :param driver: The Selenium WebDriver instance.
    :param folder_path: The folder where the screenshot will be saved.
    :param file_name: The name of the screenshot file.
    """
    try:
        # Ensure the folder exists, create it if it does not
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        # Define the full path to save the screenshot
        screenshot_path = os.path.join(folder_path, file_name)

        # Save the screenshot
        driver.execute_script("document.body.style.zoom='80%'")
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*'))
        )
        driver.save_screenshot(screenshot_path)
        
        driver.execute_script("document.body.style.zoom='100%'")
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*'))
        )
        print(f'Screenshot saved as {screenshot_path}.')

    except Exception as e:
        print(f"An error occurred while saving the screenshot: {e}")
        raise


def check_and_click_elements(driver, wait_time=10):
    try:
        # Wait for the presence of the text elements
        wait = WebDriverWait(driver, wait_time)

        # Check for the presence of the "Monitoring" text

        # If "Monitoring" is found, click the associated button

        click_element_by_id(driver, "__xmlview38--idVariantManagement-trigger-img")
        # button = driver.find_element(By.XPATH, '//span[@id="__xmlview188--idVariantManagement-trigger-inner"]')
        # button.click()
        print("Button associated with 'Monitoring' clicked.")
        time.sleep(4)

        # Check for the presence of the "All" text

        # If "All" is found, click the associated element

        # click_element_by_text(driver, "All")
        click_element_by_id(driver, "__xmlview38--idVariantManagement-trigger-item-1")
        print("Element with text 'All' clicked.")

        time.sleep(4)
        return True

    except Exception as e:
        print(f"An error occurred: {e}")
        raise


def check_last_update_successful(driver, current_day):
    try:
        is_last_update_successful = False
        is_correct_day = False
        sucText = get_text_by_partial_id(driver, "--backupDetailsStatus-text").strip()

        if sucText.lower() == "successful":
            is_last_update_successful = True

        is_correct_day = True
        if is_correct_day and is_last_update_successful:
            edit_excel(
                "./TestDB/Output.xlsx", "Daily Checks", current_day, 7, "Completed", 1
            )
        else:
            edit_excel("./TestDB/Output.xlsx", "Daily Checks", current_day, 7, "warning", 0)
        
        return is_last_update_successful
    except Exception as e:
        print('Failed to check database')
        edit_excel("./TestDB/Output.xlsx", "Daily Checks", current_day, 7, "failed",-1)
        return False
    

# ---------------------------------------------------------
def create_excel_with_table_in_folder(folder_path, file_name):
    """
    Creates an Excel file with a predefined table in the specified folder.

    Args:
    folder_path (str): The path to the folder where the Excel file should be created.
    file_name (str): The name of the Excel file to be created.
    """
    # Ensure the folder exists
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    # Define the file path
    file_path = os.path.join(folder_path, file_name)

    # Check if the file already exists
    if os.path.exists(file_path):
        print(f"The file '{file_name}' already exists in the folder '{folder_path}'.")
        return  # Exit the function to avoid overwriting the file

    # Define the data
    data = {
        'Database check': [
            'Check if service interrupted/Any inconsistencies with logging on',
            'Memory/CPU usage',
            'Disk space usage',
            'DB lock check',
            'System Dumps',
            'Backup check',
            'Certificate check'
        ],
        'Monday': ['', '', '', '', '', '', ''],
        'Tuesday': ['', '', '', '', '', '', ''],
        'Wednesday': ['', '', '', '', '', '', ''],
        'Thursday': ['', '', '', '', '', '', ''],
        'Friday': ['', '', '', '', '', '', ''],
        'Saturday': ['', '', '', '', '', '', ''],
        'Sunday': ['', '', '', '', '', '', '']
    }

    # Create a DataFrame from the data
    df = pd.DataFrame(data)

    # Write the DataFrame to an Excel file
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Daily Checks", index=False)

    print(f"Excel file '{file_path}' created successfully with the table.")
    return True

    return True


def edit_excel(file_path, sheet_name, column_name, row_number, new_text, color):
    try:
        # Load the workbook and select the sheet
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        # Find the column index for the given column name
        column_index = None
        for cell in sheet[1]:  # Assuming the first row contains headers
            if cell.value == column_name:
                column_index = cell.column
                break

        if column_index is None:
            raise ValueError(f"Column '{column_name}' not found.")

        # Access the cell to edit
        cell_to_edit = sheet.cell(row=row_number, column=column_index)

        # Update the cell value
        cell_to_edit.value = new_text

        # Set the fill color to green
        basic_color = "808080"
        if color == 1:
            basic_color = "00B050"
        elif color == -1:
            basic_color = "FF0000"
        elif color == 0:
            basic_color = "FFA500"

        green_fill = PatternFill(
            start_color=basic_color, end_color=basic_color, fill_type="solid"
        )
        cell_to_edit.fill = green_fill

        # Save the changes to the workbook
        workbook.save(file_path)
        print(f"Cell in row {row_number}, column '{column_name}' updated successfully.")
        return True

    except Exception as e:
        print(f"An error occurred: {e}")
        return False


def open_html_in_browser(html_file):
    driver = webdriver.Chrome()  # Ensure you have chromedriver installed
    html_path = "report.html"
    driver.get(html_path)


# --------------------------------------------------------------------------
from datetime import datetime


def create_html_with_buttons(filename, function_statuses):
    """
    Creates an HTML file with buttons indicating the status of each function and includes images in the buttons.

    :param filename: Name of the HTML file to be created
    :param function_statuses: Dictionary with function names as keys and their statuses as values
    """
    html_template = """
    <!DOCTYPE html>
    <html>
    <head>
        <title>Function Status</title>
        <style>
            .button {{
                padding: 10px 20px;
                margin: 5px;
                border: none;
                font-size: 16px;
                cursor: pointer;
                background-color: grey; /* default color */
                color: white;
                border-radius: 5px;
                transition: background-color 0.3s ease;
            }}
            .success {{
                background-color: green; /* Turns green when successful */
            }}
            .failure {{
                background-color: red; /* Red for failure */
            }}
            .warning {{
                background-color: #ffc107; /* Yellow for warning */
            }}
            .button a {{
                text-decoration: none;
                color: white;
            }}
        </style>
    </head>
    <body>
        <h1>Function Status Report</h1>
        <p>Generated on {timestamp}</p>
        {buttons}
        <script>
            function openImage(imagePath) {{
                window.open(imagePath, '_blank');
            }}
        </script>
    </body>
    </html>
    """

    buttons = ""
    for func_name, status in function_statuses.items():
        # Handle boolean values as well as string statuses
        if status == 1:
            color_class = "success"
        elif status == -1:
            color_class = "failure"
        elif status == 0:
            color_class = "warning"
        elif status == -2:
            color_class = "norun"
        # else:
        #     color_class = "success" if status == "success" else "failure"

        screenshot_path = ""  # Initialize to avoid UnboundLocalError

        if isinstance(status, tuple):
            status, screenshot_path = status

        # Add the button with text and link to the full-screen image
        if screenshot_path:
            buttons += f'<button class="button {color_class}" onclick="openImage(\'{screenshot_path}\')">{func_name}</button>\n'
        else:
            buttons += f'<button class="button {color_class}">{func_name}</button>\n'

    # Fill the template with the timestamp and buttons
    html_content = html_template.format(
        timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"), buttons=buttons
    )

    with open(filename, "w") as file:
        file.write(html_content)

    print(f"HTML file '{filename}' created successfully.")


#  -------------------------------------------------------------------------
images = [
    {"src": "TestDB/login.png", "alt": "Login"},
    {"src": "TestDB/databaseSelected.png", "alt": "databaseSelected.png"},
    {"src": "TestDB/Backups.png", "alt": "Backups.png"},
    {"src": "TestDB/CPU.png", "alt": "CPU.png"},
]


def insert_images_into_html(file_path, images):
    # Open and read the HTML file
    with open(file_path, "r") as file:
        html_content = file.readlines()

    # Open the file in write mode
    with open(file_path, "w") as file:
        for line in html_content:
            file.write(line)

        # Insert images dynamically
        for image in images:
            img_tag = f'<img src="{image["src"]}" alt="{image["alt"]}" style="width:100%;max-width:300px;margin:10px 0;">\n'
            file.write(img_tag)

    print("Images inserted successfully!")


def main():
    url = "https://hanahcdbdev.mud.internal.co.za:39630/"  # Replace with the URL that triggers the warning

    # Set up Chrome options
    chrome_options = Options()
    # chrome_options.add_argument("--headless")  # Optional: Run in headless mode (no GUI)
    chrome_options.add_argument(
        "--ignore-certificate-errors"
    )  # Ignore certificate errors
    chrome_options.add_argument("--start-maximized")

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.set_window_size(1920, 1580)
    current_day = get_current_day_in_words()
    create_excel_with_table_in_folder("TestDB", "Output.xlsx")

    try:
        # create_html_with_buttons("dynamic_buttons.html")
        # Open the website
        print("curr : ", current_day)

        driver.get(url)

        loginans1 = -2
        backup_ans = -2
        loginpic = -2
        backupSuccessful = -2
        Alerts = -2
        CPUans = -2
        Services = -2
        Memory = -2
        Disk = -2
        Data_Encryption = -2
        Trust_Configuration = -2

        # Bypass the security warning
        bypass_security_warning(driver)

        click_element_by_text(driver, "hana-cockpit")

        loginans1 = login(driver)
        print(loginans1)

        if not loginans1:
            raise ValueError("Failed login")

        # perform_checks(driver, ans1)

        save_screenshot(driver,"TestDB","login.png")
        print("Screenshot saved.")
        # driver.execute_script("document.body.style.zoom='80%'")
        # time.sleep(5)
        
        selectDatabase(driver,"Training", "TESTDB@DHB","TestDB")
        # selectDatabase(driver,"Training", "DHB@DHB","TestDB")

        check_and_click_elements(driver)

        # Search for backups
        search_input_field(driver, "idSearchFieldOVP-I", "backups")
        click_element_by_id(driver, "idSearchFieldOVP-search")
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*"))
        )
        
        backupSuccessful = check_last_update_successful(driver, current_day)
        print("-"*100)
        print("backups check", backupSuccessful)
        
        click_element_by_partial_id(driver,"--lastbackup")
        
        time.sleep(4)
        driver.execute_script("document.body.style.zoom='80%'")
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*'))
        )
        save_screenshot(driver,"TestDB","Backups.png")
        driver.execute_script("document.body.style.zoom='100%';")
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, '//*'))
        )
        
        click_back_icon(driver)
        time.sleep(4)

        # Search for CPU
        search_input_field(driver, "idSearchFieldOVP-I", "CPU")
        click_element_by_id(driver, "idSearchFieldOVP-search")

        time.sleep(10)
        
        save_screenshot(driver,"TestDB","CPU.png")
        print("Screenshot saved.")
        
        # Search for Alerts 
        search_input_field(driver, 'idSearchFieldOVP-I', 'Alerts')
        click_element_by_id(driver, "idSearchFieldOVP-search")

        time.sleep(10)
        
        save_screenshot(driver,"TestDB","Alerts.png")
        print("Screenshot saved.")
        
        # Search for Services 
        search_input_field(driver, 'idSearchFieldOVP-I', 'Services')
        click_element_by_id(driver, "idSearchFieldOVP-search")

        time.sleep(10)
        
        save_screenshot(driver,"TestDB","Services.png")
        print("Screenshot saved.")
        
        # Search for Memory U 
        search_input_field(driver, 'idSearchFieldOVP-I', 'Memory U')
        click_element_by_id(driver, "idSearchFieldOVP-search")

        time.sleep(10)
        
        save_screenshot(driver,"TestDB","Memory.png")
        print("Screenshot saved.")
        
        # Search for Disk
        search_input_field(driver, "idSearchFieldOVP-I", "Disk")
        click_element_by_id(driver, "idSearchFieldOVP-search")

        time.sleep(10)
        
        save_screenshot(driver,"TestDB","Disk.png")
        print("Screenshot saved.")
        
        # Search for Data Encryption
        search_input_field(driver, "idSearchFieldOVP-I", "Data En")
        click_element_by_id(driver, "idSearchFieldOVP-search")
        
        time.sleep(10)
        
        save_screenshot(driver,"TestDB","Encryption.png")
        print("Screenshot saved.")
        
        # Search for Trust Configuration - Certification
        search_input_field(driver, "idSearchFieldOVP-I", "Trust Configuration")
        click_element_by_id(driver, "idSearchFieldOVP-search")

        time.sleep(10)
        
        save_screenshot(driver,"TestDB","Certification.png")
        print("Screenshot saved.")
        
        
        

    except Exception as e:
        print(f"An error occurred: {e}")

    finally:
        # Close the browser
        function_statuses = {
            "Login": loginans1,
            "Logged in": (loginpic, "TestDB/login.png"),
            "Database_screenshot": backup_ans,
            "Back up : " + backUpText: backupSuccessful,
            "CPU": (CPU_image, "TestDB/CPU.png"),
            "Alerts": (Alerts_image, "TestDB/Alerts.png"),
            "Services": (Services_image, "TestDB/Services.png"),
            "Memory": (Services_image, "TestDB/Memory.png"),
            "Disk": (Disk_image, "TestDB/Disk.png"),
            "Data Encryption": (Data_Encryption_image, "TestDB/Encryption.png"),
            "Trust Configuration": (Certification_image, "TestDB/Certification.png"),
        }
        create_html_with_buttons("daily-checks.html", function_statuses)
        driver.quit()


if __name__ == "__main__":
    main()
